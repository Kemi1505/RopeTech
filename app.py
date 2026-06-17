import os
import pandas as pd
from flask import Flask, request, render_template, send_file, flash, redirect, url_for
from werkzeug.utils import secure_filename
from mailmerge import MailMerge
import tempfile
import traceback
import math
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
app.secret_key = 'replace-this-with-a-long-random-string'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs('templates', exist_ok=True)

# -----------------------------------------------------------------
# ALLOWED FILE CHECK
# -----------------------------------------------------------------
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'xlsx', 'xls'}

# -----------------------------------------------------------------
# EQUIPMENT CONFIGURATION FOR CERTIFICATES
# -----------------------------------------------------------------
EQUIPMENT_CONFIG = {
    # Shackle types – each has its own template and uses the same mapping file
    'Bow Shackle - Bolt & Nut': {
        'template': 'templates/Bow Shackle - Bolt Template.docx',
        'sheet': 'Bow Shackles | BN',
        'mapping_file': 'shackles/shackle-data.xlsx'
    },
    'Bow Shackle Screw Pin': {
        'template': 'templates/Bow Shackle - Screw Template.docx',
        'sheet': 'Bow Shackles | SP',
        'mapping_file': 'shackles/shackle-data.xlsx'
    },
    'Dee Shackle Bolt & Nut': {
        'template': 'templates/Dee Shackle - Bolt Template.docx',
        'sheet': 'Dee Shackles | BN',
        'mapping_file': 'shackles/shackle-data.xlsx'
    },
    'Dee Shackle Screw Pin': {
        'template': 'templates/Dee Shackle - Screw Template.docx',
        'sheet': 'Dee Shackles | SP',
        'mapping_file': 'shackles/shackle-data.xlsx'
    },
    # Other equipment – no mapping
    'Body Harness': {
        'template': 'templates/Body Harness Belt Template.docx',
        'sheet': 'Body Harness'
    },
    'Lanyard': {
        'template': 'templates/Lanyard Template.docx',
        'sheet': 'Lanyard'
    },
    'Flat Webbing Belts': {
        'template': 'templates/Flat Webbing Belts Template.docx',
        'sheet': 'Flat Webbing Belts'
    },
    'Chain Blocks': {
        'template': 'templates/Chain Blocks Template.docx',
        'sheet': 'Chain Blocks'
    },
    'Lever Hoists': {
        'template': 'templates/Lever Hoist Template.docx',
        'sheet': 'Lever Hoists'
    },
    'Plate Clamps': {
        'template': 'templates/Plate Clamp Template.docx',
        'sheet': 'Plate Clamps'
    },
    'Wire Ropes': {
        'template': 'templates/Wire Ropes Template.docx',
        'sheet': 'Wire Ropes'
    }
}

# -----------------------------------------------------------------
# FORMAT VALUE (remove .0 from whole numbers)
# -----------------------------------------------------------------
def format_value(v):
    if pd.isna(v) or v is None:
        return ''
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        else:
            return str(v)
    return str(v)

# -----------------------------------------------------------------
# LOAD MAPPING (for shackles)
# -----------------------------------------------------------------
def load_mapping(mapping_file):
    df = pd.read_excel(mapping_file)
    df = df.where(pd.notnull(df), None)
    mapping = {}
    for _, row in df.iterrows():
        try:
            load = float(row.iloc[0])
        except:
            continue
        formatted_row = {col: format_value(val) for col, val in row.items()}
        mapping[load] = formatted_row
    return mapping

# -----------------------------------------------------------------
# CERTIFICATE GENERATION (mail merge)
# -----------------------------------------------------------------
def generate_certificates(excel_path, template_path, sheet_name, output_filename, mapping_file=None):
    df = pd.read_excel(excel_path, sheet_name=sheet_name)
    df = df.where(pd.notnull(df), None)
    raw_records = df.to_dict(orient='records')
    if not raw_records:
        raise ValueError("No data rows in selected sheet.")
    
    # Load mapping if needed
    mapping = load_mapping(mapping_file) if mapping_file else None
    
    merge_records = []
    for rec in raw_records:
        load_val = rec.get('WLL')
        try:
            load = float(load_val) if load_val is not None else 0.0
        except:
            load = 0.0
        
        formatted_rec = {k: format_value(v) for k, v in rec.items()}
        
        if mapping:
            mapping_vals = mapping.get(load, mapping.get(list(mapping.keys())[0], {}))
            formatted_rec.update(mapping_vals)
        
        merge_records.append(formatted_rec)
    
    output_dir = tempfile.mkdtemp()
    output_path = os.path.join(output_dir, output_filename)
    with MailMerge(template_path) as doc:
        doc.merge_templates(merge_records, separator='page_break')
        doc.write(output_path)
    return output_path

# -----------------------------------------------------------------
# REGISTER GENERATION (Excel)
# -----------------------------------------------------------------

# Column mapping for each equipment type (for register)
EQUIPMENT_COLUMNS = {
    'Shackles': {
        'columns': ['S/N',	'EQUIPMENT', 'ID NO', 'MAKE', 'NORMINAL SIZE', 'TYPE', 'WLL', 'DATE INSPECTED', 'DUE DATE', 'LOCATION', 'REMARK'],
        'sheet_names': ['Bow Shackles | BN', 'Bow Shackles | SP', 'Dee Shackles | BN', 'Dee Shackles | SP'],
        'group_as': 'SHACKLES'
    },
    'Flat Webbing Belts': {
        'columns': ['S/N',	'EQUIPMENT', 'MAKER', 'SERIAL NO', 'LENGTH', 'SWL', 'COLOUR', 'DATE INSPECTED', 'DUE DATE', 'LOCATION', 'REMARK']
    },
    'Full Body Harness': {
        'columns': ['S/N', 'EQUIPMENT', 'MAKER', 'ID NO', 'MODEL NO', 'LENGTH', 'SWL', 'DATE INSPECTED', 'DUE DATE', 'LOCATION', 'REMARK']
    },
    'Chain Blocks': {
        'columns': ['S/N', 'EQUIPMENT', 'MAKER', 'SERIAL NO', 'MODEL NO', 'SWL', 'DIAMETER', 'DATE INSPECTED', 'DUE DATE', 'LOCATION', 'REMARK']
    },
    'Lever Hoists': {
        'columns': ['S/N', 'EQUIPMENT', 'MAKER', 'SERIAL NO', 'MODEL NO', 'SWL', 'DIAMETER', 'DATE INSPECTED', 'DUE DATE', 'LOCATION', 'REMARK']
    },
    'Wire Ropes': {
        'columns': ['S/N', 'EQUIPMENT', 'ID NO', 'LENGTH', 'DIAMETER', 'TERMINATION', 'SWL', 'DATE INSPECTED', 'DUE DATE', 'LOCATION', 'REMARK']
    },
    # Add more equipment types as needed       
}

SHACKLE_MAPPING_FILE = "shackles/shackle-data.xlsx"

def load_shackle_mapping(mapping_file):
    if not os.path.exists(mapping_file):
        raise FileNotFoundError(f"Shackle mapping file not found: {mapping_file}")
    df = pd.read_excel(mapping_file)
    df = df.fillna('')
    mapping = {}
    for _, row in df.iterrows():
        equipment = str(row.get('Equipment', '')).strip()
        type_ = str(row.get('Type', '')).strip()
        key = f"{equipment}|{type_}".lower()
        mapping[key] = row.to_dict()
    return mapping

def parse_shackle_sheet_name(sheet_name):
    parts = sheet_name.split('|')
    if len(parts) == 2:
        equipment = parts[0].strip()
        type_code = parts[1].strip().upper()
        if type_code == 'BN':
            type_ = 'Bolt & Nut'
        elif type_code == 'SP':
            type_ = 'Screw Pin'
        else:
            type_ = type_code
        return equipment, type_
    return sheet_name, None

def get_nominal_size(mapping_vals):
    w = str(mapping_vals.get('W', '')).strip()
    n1 = str(mapping_vals.get('N1', '')).strip()
    n2 = str(mapping_vals.get('N2', '')).strip()
    
    parts = []
    if w:
        parts.append(w)
    if n1 and n2:
        parts.append(f"{n1}/{n2}")
    elif n1 and not n2:
        parts.append(n1)
    
    if parts:
        return " ".join(parts) + '"'
    else:
        return ''

def generate_register(excel_path, output_filename):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # --- Use load_mapping (keyed by WLL) instead of load_shackle_mapping ---
    mapping_by_wll = load_mapping(SHACKLE_MAPPING_FILE)   # returns dict {WLL: row_dict}
    excel_file = pd.ExcelFile(excel_path)
    sheet_names = excel_file.sheet_names
    excel_file.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment Register"

    # Styles
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    arial9 = Font(name='Arial', size=9)
    arial9_bold = Font(name='Arial', size=9, bold=True)
    arial11_bold = Font(name='Arial', size=11, bold=True)   # <-- new
    center_align = Alignment(horizontal='center', vertical='center')

    def resolve_column(data_headers, target_col):
        # ... (keep your existing resolve_column function) ...
        if not data_headers:
            return None
        target_lower = target_col.lower().strip()
        for h in data_headers:
            if h.lower().strip() == target_lower:
                return h
        variations = {
            'id no': ['id no', 'id', 'serial no', 'serial', 'identification'],
            'maker': ['maker', 'make', 'manufacturer'],
            'swl': ['swl', 'wll', 'safe working load', 'working load limit'],
            'wll': ['wll', 'swl', 'safe working load', 'working load limit'],
            'date inspected': ['date inspected', 'inspection date', 'date1'],
            'due date': ['due date', 'expiry date', 'date2'],
            'equipment': ['equipment', 'item', 'description'],
            'norminal size': ['norminal size', 'nominal size', 'size'],
            'serial no': ['serial no', 'serial', 'id no', 'id'],
            'model no': ['model no', 'model', 'part no'],
            'length': ['length', 'overall length'],
            'diameter': ['diameter', 'dia'],
            'termination': ['termination', 'end fitting'],
            'colour': ['colour', 'color'],
        }
        for key, variants in variations.items():
            if key in target_lower:
                for v in variants:
                    for h in data_headers:
                        if v in h.lower().strip():
                            return h
        return None

    current_row = 11

    for sheet_name in sheet_names:
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
        df = df.fillna('').astype(str)
        if df.empty:
            continue

        data_headers = df.columns.tolist()

        # Determine equipment name and output column list
        equipment_name = None
        output_columns = None
        is_shackle = False
        shackle_equipment = None
        shackle_type = None

        if any(sheet_name.startswith(name) for name in ['Bow Shackles', 'Dee Shackles']):
            is_shackle = True
            equipment_name = 'SHACKLES'
            shackle_equipment, shackle_type = parse_shackle_sheet_name(sheet_name)
            output_columns = EQUIPMENT_COLUMNS['Shackles']['columns']
        else:
            matched = False
            for equip_name, config in EQUIPMENT_COLUMNS.items():
                if equip_name != 'Shackles' and sheet_name.lower() == equip_name.lower():
                    equipment_name = sheet_name.upper()
                    output_columns = config['columns']
                    matched = True
                    break
            if not matched:
                equipment_name = sheet_name.upper()
                output_columns = data_headers

        # --- Equipment header row ---
        for col_idx, col_name in enumerate(output_columns, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=equipment_name)
            cell.font = arial9_bold
            cell.font = arial11_bold  
            cell.alignment = center_align
            cell.fill = PatternFill(start_color="7F7F7F", end_color="7F7F7F", fill_type="solid")
            cell.border = thin_border
        if len(output_columns) > 0:
            ws.merge_cells(start_row=current_row, start_column=1,
                          end_row=current_row, end_column=min(len(output_columns), 11))
        current_row += 1

        # --- Column headers ---
        for col_idx, col_name in enumerate(output_columns, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=col_name)
            cell.font = arial9_bold
            cell.alignment = center_align
            cell.border = thin_border
        current_row += 1

        # --- Data rows ---
        data_rows = df.to_dict(orient='records')
        sn = 1
        for row_data in data_rows:
            row_idx = current_row
            ws.cell(row=row_idx, column=1, value=sn).font = arial9
            ws.cell(row=row_idx, column=1, value=sn).alignment = center_align
            ws.cell(row=row_idx, column=1, value=sn).border = thin_border

            # --- For shackles, fetch mapping by WLL from this row ---
            row_mapping_vals = {}
            if is_shackle:
                wll_from_data = row_data.get('WLL', '') or row_data.get('SWL', '')
                if wll_from_data:
                    try:
                        wll_float = float(str(wll_from_data).strip())
                        row_mapping_vals = mapping_by_wll.get(wll_float, {})   # <-- FIX: use mapping_by_wll
                    except (ValueError, TypeError):
                        pass

            for col_idx, output_col in enumerate(output_columns, start=1):
                if col_idx == 1:
                    continue
                value = ''

                if is_shackle:
                    if output_col == 'EQUIPMENT' and shackle_equipment:
                        # --- NEW: singularize ---
                        singular_map = {
                            'Bow Shackles': 'Bow Shackle',
                            'Dee Shackles': 'Dee Shackle'
                        }
                        value = singular_map.get(shackle_equipment, shackle_equipment)
                    elif output_col == 'TYPE' and shackle_type:
                        value = shackle_type
                    elif output_col == 'NORMINAL SIZE':
                        value = get_nominal_size(row_mapping_vals)
                    elif output_col == 'REMARK':
                        value = 'Ok for use'
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.font = arial9_bold
                        cell.alignment = center_align
                        cell.border = thin_border
                        continue
                    else:
                        source_col = resolve_column(data_headers, output_col)
                        if source_col:
                            value = row_data.get(source_col, '')
                else:
                    # EQUIPMENT with fallback
                    if output_col == 'EQUIPMENT':
                        source_col = resolve_column(data_headers, output_col)
                        if source_col:
                            value = row_data.get(source_col, '')
                        else:
                            desc_col = resolve_column(data_headers, 'DESCRIPTION')
                            if desc_col:
                                value = row_data.get(desc_col, '')
                            else:
                                value = ''
                        if not value or value.strip() == '':
                            hardcoded_equipment = {
                                'CHAIN BLOCKS': 'Chain Block',
                                'LEVER HOISTS': 'Lever Hoist',
                                'FLAT WEBBING BELTS': 'Flat Webbing Belt',
                            }
                            value = hardcoded_equipment.get(equipment_name, equipment_name)
                    elif output_col == 'REMARK':
                        value = 'Ok for use'
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.font = arial9_bold
                        cell.alignment = center_align
                        cell.border = thin_border
                        continue
                    else:
                        source_col = resolve_column(data_headers, output_col)
                        if source_col:
                            value = row_data.get(source_col, '')

                # --- Special formatting for SWL/WLL ---
                if output_col.upper() in ['SWL', 'WLL'] and value:
                    value = str(value).strip()
                    if not value.endswith('T') and value.replace('.', '').replace('-', '').isdigit():
                        value = value + 'T'

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = arial9
                cell.alignment = center_align
                cell.border = thin_border

            current_row += 1
            sn += 1

        #current_row += 1   # blank line between groups

    output_dir = tempfile.mkdtemp()
    output_path = os.path.join(output_dir, output_filename)
    wb.save(output_path)
    return output_path
# -----------------------------------------------------------------
# FLASK ROUTES
# -----------------------------------------------------------------
@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        equipment = request.form.get('equipment')
        output_name = request.form.get('output_name', '').strip()
        if not output_name:
            flash('Please enter an output file name.')
            return redirect(url_for('index'))
        if not output_name.endswith('.docx'):
            output_name += '.docx'

        if 'excel_file' not in request.files:
            flash('No file uploaded.')
            return redirect(url_for('index'))
        file = request.files['excel_file']
        if file.filename == '':
            flash('No file selected.')
            return redirect(url_for('index'))
        if not allowed_file(file.filename):
            flash('Only .xlsx or .xls files are allowed.')
            return redirect(url_for('index'))
        if equipment not in EQUIPMENT_CONFIG:
            flash('Invalid equipment selected.')
            return redirect(url_for('index'))

        excel_filename = secure_filename(file.filename)
        excel_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_filename)
        file.save(excel_path)

        try:
            config = EQUIPMENT_CONFIG[equipment]
            template_path = config['template']
            sheet_name = config['sheet']
            mapping_file = config.get('mapping_file')
            output_path = generate_certificates(excel_path, template_path, sheet_name, output_name, mapping_file)
            return send_file(
                output_path,
                as_attachment=True,
                download_name=output_name,
                mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
            )
        except Exception as e:
            traceback.print_exc()
            flash(str(e))
            return redirect(url_for('index'))
        finally:
            if os.path.exists(excel_path):
                os.remove(excel_path)

    return render_template('index.html', equipment_list=EQUIPMENT_CONFIG.keys())

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        output_name = request.form.get('output_name', '').strip()
        if not output_name:
            flash('Please enter an output file name.')
            return redirect(url_for('register'))
        if not output_name.endswith('.xlsx'):
            output_name += '.xlsx'

        if 'excel_file' not in request.files:
            flash('No file uploaded.')
            return redirect(url_for('register'))
        file = request.files['excel_file']
        if file.filename == '':
            flash('No file selected.')
            return redirect(url_for('register'))
        if not allowed_file(file.filename):
            flash('Only .xlsx or .xls files are allowed.')
            return redirect(url_for('register'))

        excel_filename = secure_filename(file.filename)
        excel_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_filename)
        file.save(excel_path)

        try:
            output_path = generate_register(excel_path, output_name)
            return send_file(
                output_path,
                as_attachment=True,
                download_name=output_name,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        except Exception as e:
            traceback.print_exc()
            flash(str(e))
            return redirect(url_for('register'))
        finally:
            if os.path.exists(excel_path):
                os.remove(excel_path)

    return render_template('register.html')

# -----------------------------------------------------------------
# RUN THE APP
#if __name__ == '__main__':
#   app.run(debug=True, host='0.0.0.0', port=7000)
# -----------------------------------------------------------------

# if __name__ == '__main__':
#     from waitress import serve
#     serve(app, host='0.0.0.0', port=5000)

if __name__ == '__main__':
   app.run(debug=True, host='0.0.0.0', port=7000)